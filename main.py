from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
from os import listdir
import requests

# Hello! I'm Weslley Tavares and I'm a 14-year-old junior programmer.
# I've created this small webscraping project for my GitHub.
# Feel free to test, evaluate, or study the code.

# Use "pip install -r requirements.txt" to install openpyxl & bs4
# Site "https://books.toscrape.com/" was utilized, designed for web scraping.


class BookScraper:
    """Class with the functions to perform webscrapping"""

    def __init__(self):
        self.spreadsheet = self.__setup_spreadsheet()
        self.sheet = self.spreadsheet['Sheet']

    def __setup_spreadsheet(self):
        """Set up spreadsheet"""
        # Checking if the "booklist.xlsx" file exists
        if 'booklist.xlsx' in listdir():
            # Returns the spreadsheet if it exists
            return load_workbook('booklist.xlsx')
        else:
            # Creates a new spreadsheet and returns it
            spreadsheet = Workbook()
            spreadsheet['Sheet'].append(['Book', 'Price', 'Availability'])
            return spreadsheet

    def scrape_books(self, page: int = 1):
        """_summary_
        Args:
            page (int, optional): Page to be viewed. Defaults to 1.
        """
        # Getting page
        link = f'https://books.toscrape.com/catalogue/page-{page}.html'
        response = requests.get(link)

        # Checks if the request was a success
        if response.ok:
            # Setting up web scraping
            html = response.text
            scrapper = BeautifulSoup(html, 'html.parser')

            # Book list
            book_list = scrapper.find_all(
                'article', attrs={'class': 'product_pod'})

            # Book info
            for book in book_list:
                # Book info
                book_name = book.find('h3').a['title']
                book_price = book.find(
                    'p', attrs={'class': 'price_color'}).text
                book_avail = book.find(
                    'p', attrs={'class': 'instock availability'}).text.strip()

                # Checks if the book is already in the spreadsheet
                book_coord = None
                for sheet_rows in self.sheet.iter_rows(min_row=2):
                    # Breaks the loop if you've found a coordinate
                    if book_coord is not None:
                        # Updating book info
                        book_pos = book_coord[1:]
                        self.sheet['B'+book_pos] = book_price[1:]
                        self.sheet['C'+book_pos] = book_avail
                        break

                    # Going into every row
                    for row in sheet_rows:
                        # Breaks again if passed name section or empty name
                        if 'A' not in row.coordinate or row.value is None:
                            break

                        # Checks if (row.value == book_name) and updates book_cord
                        if row.value == book_name:
                            book_coord = row.coordinate
                            break

                # If the book doesn't exist, create a new one.
                if book_coord is None:
                    self.sheet.append(
                        [book_name, book_price[1:], book_avail]
                    )
        else:
            raise ValueError('Please select a page between 1 and 50.')

    def save_spreadsheet(self):
        """_summary_
        Saves the spreedsheet as a file. (booklist.xlsx)
        """
        self.spreadsheet.save('booklist.xlsx')
        print(f'Spreadsheet saved as "booklist.xlsx" in {os.path.pardir}')


if __name__ == "__main__":
    book_scraper = BookScraper()
    book_scraper.scrape_books(int(input('Insert page (1-50): ')))
    book_scraper.save_spreadsheet()
